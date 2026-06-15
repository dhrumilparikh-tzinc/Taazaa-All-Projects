"""Dev utility — debug XLSX/CSV file processing step by step."""

import sys

sys.path.insert(
    0, r"C:\Users\Dhrumil.parikh\OneDrive - Taazaa Tech Pvt Ltd\Desktop\playbook_final\geminirag"
)

from sqlalchemy import text

from app.models.db import get_engine

# Get one of the problem XLSX files
engine = get_engine()
with engine.connect() as conn:
    rows = conn.execute(
        text(
            "SELECT filename, file_path FROM jobs WHERE filename IN ('SUM-2025-3241.xlsx', 'INV-2025-4291-GT.xlsx', 'CTR-2026-3042.xlsx') LIMIT 3"
        )
    ).fetchall()

for r in rows:
    print(f"\n=== {r[0]} ===")
    try:
        import openpyxl

        wb = openpyxl.load_workbook(r[1], read_only=True, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for sheet_name in wb.sheetnames[:1]:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows())
            print(
                f"Sheet '{sheet_name}': {len(all_rows)} rows, {len(list(ws.iter_cols())) if all_rows else 0} cols"
            )
    except Exception as e:
        print(f"  Error: {e}")

# Build the prompt for one file and measure it
print("\n--- Testing SUM file text length ---")
from app.processors.xlsx_proc import XLSXProcessor


class MockJob:
    def __init__(self, path, filename):
        self.file_path = path
        self.filename = filename
        self.id = "test"
        self.user_id = "test"
        self.file_type = "xlsx"


class MockSettings:
    GROQ_API_KEY = "test"
    GROQ_MODEL = "test"
    GROQ_VISION_MODEL = "test"


for r in rows[:1]:
    job = MockJob(r[1], r[0])
    proc = XLSXProcessor(job=job, settings=MockSettings())
    text = proc.extract()
    print(f"Full extracted text length: {len(text)} chars")
    print(f"After 1500-char truncation: {min(len(text), 1500)} chars")
    trunc = text[:1500]

    # Build actual prompt
    prompt = f"""You are a data analyst. Analyse the following spreadsheet data and return ONLY valid JSON.
No preamble, no markdown code blocks, just raw JSON.

Return this exact structure:
{{
  "title": "spreadsheet title or filename",
  "summary": "2-3 sentence summary of what this data contains",
  "sheets": ["list of sheet names found"],
  "column_descriptions": {{"column_name": "what it likely represents"}},
  "key_insights": ["notable patterns, max/min values, trends noticed"],
  "row_count": 0
}}

Spreadsheet data:
{trunc}
"""
    print(f"Full prompt char length: {len(prompt)}")
    print(f"Prompt first 300 chars:")
    print(prompt[:300])
    print(f"\nLast 300 chars of prompt:")
    print(prompt[-300:])
