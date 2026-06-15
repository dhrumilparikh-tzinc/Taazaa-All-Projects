"""
XLSX and CSV file processor.

Iterates every sheet in the workbook with openpyxl.  Each sheet becomes a
'## Sheet Name' markdown section containing a markdown table with column
headers.  Rows beyond _MAX_ROWS (500) are truncated to keep token budgets
manageable.  CSV files are parsed with the built-in csv module and treated as
a single-sheet workbook.

The LLM summary (title, summary, sheets, column_descriptions, key_insights,
row_count) is produced by a single Groq call after extraction.
"""

import csv

import openpyxl

from app.processors.base import BaseProcessor

_MAX_ROWS = 500  # hard cap per sheet — keeps chunks manageable


class XLSXProcessor(BaseProcessor):
    def extract(self) -> str:
        if self.job.file_path.endswith(".csv"):
            return self._extract_csv()
        return self._extract_xlsx()

    def _extract_csv(self) -> str:
        with open(self.job.file_path, newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))

        if len(rows) > _MAX_ROWS + 1:
            self.log.warning("csv_rows_truncated", total=len(rows), limit=_MAX_ROWS)
            rows = rows[: _MAX_ROWS + 1]

        parts = [f"# {self.job.filename}", self._table_to_markdown(rows)]
        return "\n\n".join(parts)

    def _extract_xlsx(self) -> str:
        wb = openpyxl.load_workbook(self.job.file_path, read_only=True, data_only=True)
        parts = [f"# {self.job.filename}"]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [
                [str(cell.value) if cell.value is not None else "" for cell in row]
                for row in ws.iter_rows()
            ]

            if len(rows) > _MAX_ROWS + 1:
                self.log.warning(
                    "xlsx_rows_truncated",
                    sheet=sheet_name,
                    total=len(rows),
                    limit=_MAX_ROWS,
                )
                rows = rows[: _MAX_ROWS + 1]

            parts.append(f"## Sheet: {sheet_name}")
            parts.append(self._table_to_markdown(rows))

        return "\n\n".join(parts)

    def summarise(self, text: str, db) -> dict:
        # Use a larger excerpt for summary since spreadsheets can have many columns
        summary_text = text[:3000] if len(text) > 3000 else text

        prompt = f"""You are a data analyst. Analyse the following spreadsheet data and return ONLY valid JSON.
No preamble, no markdown code blocks, just raw JSON.

Return this exact structure:
{{
  "title": "spreadsheet title or filename",
  "summary": "2-3 sentence summary of what this data contains",
  "sheets": ["list of sheet names found"],
  "column_descriptions": {{"column_name": "what it likely represents"}},
  "key_insights": ["notable patterns or values noticed"],
  "row_count": 0
}}

Spreadsheet data:
{summary_text}
"""
        return self._call_gemini_json(prompt, db)
